import threading
from tkinter import ttk
import customtkinter
from PIL import Image
import pandas as pd
import ctypes
import os
import sys
from datetime import datetime
import time
from openpyxl import load_workbook, Workbook
import random
import pywhatkit
import phonenumbers
from tkinter import messagebox
from tkinter import filedialog
import tkinter as tk

# =========================
# FIXED BACKEND
# =========================

class Auto_Whatsapp:
    def __init__(self, msg="", people=""):
        self.msg = msg
        self.target = people

    def ensure_logged_in(self):
        """Placeholder for logging check if using Selenium, 
        pywhatkit uses the logged-in browser session."""
        pass

    def quit_from_process(self):
        """Placeholder for cleanup."""
        pass

    def normalize_phone(self, phone):
        """Return phone number in international format without '+'"""
        try:
            # Handle scientific notation or floats from Excel
            phone_str = str(int(float(phone))) if isinstance(phone, (float, int)) else str(phone)
            x = phonenumbers.parse(f"+{phone_str.strip('+')}", None)
            if phonenumbers.is_valid_number(x):
                return str(x.country_code) + str(x.national_number)
            return None
        except Exception:
            return None

    def send_message(self, phone, message, wait_time=15):
        """
        Send message instantly to a phone number.
        """
        try:
            # Change 'phone' to 'phone_no'
            pywhatkit.sendwhatmsg_instantly(
                phone_no=f"+{phone}", 
                message=message, 
                wait_time=wait_time, 
                tab_close=True, 
                close_time=3
            )
            return "SENT ✅"
        except Exception as e:
            print("Error sending message:", e)
            return "FAILED ❌"


# =========================
# GUI PART
# =========================

class WhatsAppGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("WhatsApp Bulk Sender")
        self.file_path = None


        # Header Frame and it's labels and customization
        folder_path = "whatsapp_profile"
        folder_exists = os.path.exists(folder_path)
        status_text = ""
        status_color = ""
        border_color = ""
        fg_color_status = ""
        if folder_exists:
            status_text = "● WhatsApp Connected"
            status_color = "#4ADE80"
            border_color = "#166534"
            fg_color_status = "#14532D"
        else:
            status_text = "● WhatsApp Not Connected"
            status_color = "#94A3B8"  # Gray
            border_color = "#334155"
            fg_color_status = "#1E293B"
        
        

        header_frame = customtkinter.CTkFrame(root,fg_color="#111418", border_width=1,border_color="#1E293B")
        header_frame.pack(side='top', fill='x')
        
        left_header_frame = customtkinter.CTkFrame(header_frame,fg_color="transparent")
        left_header_frame.pack(side='left',padx=1, pady=1)

        firstLabel_font = customtkinter.CTkFont(family="Inter", size=24, weight="bold")

        customtkinter.CTkLabel(left_header_frame, text_color="#FFFFFF",text="New Campaign",font=firstLabel_font).grid(row=0,column=1, sticky='W', padx=32, pady=(20,0))
        
        secondLabel_font = customtkinter.CTkFont(family="Inter", size=14, weight="normal")
        
        customtkinter.CTkLabel(left_header_frame, text_color="#9DABB9", text="Configure and launch your bulk messaging campaign.", font=secondLabel_font).grid(row=1,column=1, sticky='W', padx=32, pady=(1,20))
        

        status_font = customtkinter.CTkFont("Inter", size=14)

        status_badge = customtkinter.CTkButton(
            header_frame, 
            text=status_text,
            text_color=status_color,
            font=status_font,
            fg_color=fg_color_status,
            border_width=1,
            border_color=border_color,
            corner_radius=20,
            width=180,
            height=32,
            hover=False,           # Disable the hover effect
            command=None,          # Ensure clicking does nothing
            cursor="arrow"         # Keep the arrow cursor so it doesn't look like a button
        )
        status_badge.pack(side='right', padx=32, pady=(10, 0))

        # Main Frame and it's customization
        self.main_frame = customtkinter.CTkFrame(root, fg_color="#101922", border_width=1,border_color="#1E293B")
        self.main_frame.pack(fill='both', expand=True)
        
        self.scrollbar = customtkinter.CTkScrollableFrame(self.main_frame, fg_color="#101922", border_width=1, border_color="#1E293B")
        self.scrollbar.pack(fill='both', expand=True, padx=10, pady=10)

        # 1. THE DASHED UPLOAD AREA (The "Inbox")
        # We use a frame as a container to hold the canvas and the button together
        #self.upload_container = customtkinter.CTkFrame(self.main_frame, fg_color="transparent", #width=963, height=201)
        #self.upload_container.pack(pady=(20, 10))


        #self.canvas = tk.Canvas(self.upload_container, bg="#101922", highlightthickness=0, width=970, #height=201)
        #self.canvas.place(x=0, y=0)
        #self.canvas.create_rectangle(0, 0, 960, 198, outline="#1E293B", width=2, dash=(10, 5))

        # Load and resize the image (e.g., to 20x20 pixels)
        image = Image.open("images/logo_excel.png")
        
        photo_image = customtkinter.CTkImage(light_image=image, size=(50,50))

        upload_btn_font = customtkinter.CTkFont(family="Inter", size=18, weight="bold")

        self.upload_btn = customtkinter.CTkButton(
            self.scrollbar, text="Upload Excel File\nDrag & drop or click to upload", command=self.upload_file, width=963, height=201,fg_color="#1C2127",image=photo_image, compound='top',font=upload_btn_font
        )
        self.upload_btn.image = photo_image # Keep reference
        self.upload_btn.pack(fill="x", padx=40, pady=20)

        # 2. THE FILE DETAILS CARD (Initially Hidden)
        self.file_card = customtkinter.CTkFrame(self.scrollbar, fg_color="#1C2127", border_width=1, border_color="#334155", height=80)
        # We don't pack it yet

        self.file_icon_label = customtkinter.CTkLabel(self.file_card, text="📄", font=("Inter", 24), text_color="#4ADE80")
        self.file_icon_label.pack(side="left", padx=20)

        self.file_details_inner = customtkinter.CTkFrame(self.file_card, fg_color="transparent")
        self.file_details_inner.pack(side="left", fill="both", expand=True, pady=10)

        self.file_name_label = customtkinter.CTkLabel(self.file_details_inner, text="", font=("Inter", 14, "bold"), text_color="#FFFFFF", anchor="w")
        self.file_name_label.pack(fill="x")

        self.file_stats_label = customtkinter.CTkLabel(self.file_details_inner, text="", font=("Inter", 12), text_color="#9DABB9", anchor="w")
        self.file_stats_label.pack(fill="x")

        self.delete_btn = customtkinter.CTkButton(self.file_card, text="🗑", width=40, fg_color="transparent", hover_color="#EF4444", command=self.remove_file)
        self.delete_btn.pack(side="right", padx=15)

        self.make_file_card_clickable()


        # Delay Message section
        
        self.delay_message = customtkinter.CTkFrame(self.scrollbar, fg_color="#1C2127", border_color="#334155", corner_radius=24)
        self.delay_message.pack(fill="x", padx=40, pady=50)

        self.delay_message.grid_columnconfigure(0, weight=1)
        font_message_delay_h1 = customtkinter.CTkFont("Inter", 18, weight="bold")

        self.message_delay_h1 = customtkinter.CTkLabel(self.delay_message,font=font_message_delay_h1,text_color="#FFFFFF", text="Message Settings")
        self.message_delay_h1.grid(row=0, column=0, sticky="w",
                           padx=24, pady=(24,14))

        font_header_delay_message = customtkinter.CTkFont("Inter", 14)

        self.header_delay_message = customtkinter.CTkLabel(self.delay_message, font=font_header_delay_message, text_color="#CBD5E1", text="Delay Time (seconds)")
        self.header_delay_message.grid(row=1, column=0,
                               sticky="w",
                               padx=24, pady=(0,8))

        font_delay_input = customtkinter.CTkFont("Inter", 16)

        # 1. Load the image using CTkImage (better for high-DPI scaling)
        image_path = "images/wall-clock.png"
        clock_image = customtkinter.CTkImage(light_image=Image.open(image_path),
                                            dark_image=Image.open(image_path),
                                            size=(32, 32)) # Adjust size to fit nicely

        # 2. Create a container Frame that acts as the "Entry Box"
        self.entry_container = customtkinter.CTkFrame(
            self.delay_message, 
            height=50, 
            fg_color="#111418", # Same as your original entry color
            border_color="#475569", 
            border_width=2, 
            corner_radius=8
        )
        self.entry_container.grid(row=2, column=0, sticky="ew", padx=24, pady=(0, 8))
        self.entry_container.grid_columnconfigure(1, weight=1) # Entry expands, icon stays

        # 3. Add the Image as a Label inside the container
        self.icon_label = customtkinter.CTkLabel(
            self.entry_container, 
            text="", 
            image=clock_image,
            compound='left'
        )
        self.icon_label.grid(row=0, column=0, padx=(12, 8), pady=10) # Left padding for the icon

        # 4. Add the Entry (Remove its borders and match background)
        self.delay_input = customtkinter.CTkEntry(
            self.entry_container, 
            height=48, # Slightly smaller than container to prevent clipping
            fg_color="transparent", 
            border_width=0, 
            text_color="#FFFFFF", 
            font=font_delay_input,
            placeholder_text="0" # Optional: looks cleaner
        )
        self.delay_input.grid(row=0, column=1, sticky="ew", padx=(0, 10))
        
        font_message_delay_info = customtkinter.CTkFont("Inter", size=12)

        self.message_delay_info = customtkinter.CTkLabel(self.delay_message, font=font_message_delay_info, text_color="#94A3B8", text="Randomized delay helps prevent spam detection.")
        self.message_delay_info.grid(row=3, column=0,  sticky="w",
                               padx=24, pady=(0,24))



        self.progress_container = customtkinter.CTkFrame(self.scrollbar, fg_color="#1C2127", border_color="#334155", border_width=1, corner_radius=16)
        # Initially hidden, just like your old progress bar
        # self.progress_container.pack(fill="x", padx=40, pady=20) 

        # Top Section: Header and Percentage
        self.prog_header_frame = customtkinter.CTkFrame(self.progress_container, fg_color="transparent")
        self.prog_header_frame.pack(fill="x", padx=24, pady=(20, 10))

        self.prog_title = customtkinter.CTkLabel(self.prog_header_frame, text="Sending in progress...", font=("Inter", 14, "bold"), text_color="#FFFFFF")
        self.prog_title.pack(side="left")

        self.percent_label = customtkinter.CTkLabel(self.prog_header_frame, text="0%", font=("Inter", 14, "bold"), text_color="#137FEC")
        self.percent_label.pack(side="right")

        # Stats Sub-header: "Messages Sent: 0 of 0"
        self.counter_label = customtkinter.CTkLabel(self.progress_container, text="Messages Sent: 0 of 0", font=("Inter", 12), text_color="#94A3B8")
        self.counter_label.pack(anchor="w", padx=24)

        # The Actual Progress Bar
        self.progressbar = customtkinter.CTkProgressBar(self.progress_container, height=10, fg_color="#111418", progress_color="#137FEC")
        self.progressbar.set(0)
        self.progressbar.pack(fill="x", padx=24, pady=15)

        # Bottom Section: Three Column Stats
        self.stats_frame = customtkinter.CTkFrame(self.progress_container, fg_color="transparent")
        self.stats_frame.pack(fill="x", padx=24, pady=(0, 20))
        self.stats_frame.grid_columnconfigure((0, 1, 2), weight=1)

        # Success Column
        self.success_val = customtkinter.CTkLabel(self.stats_frame, text="0", font=("Inter", 24, "bold"), text_color="#4ADE80")
        self.success_val.grid(row=0, column=0)
        customtkinter.CTkLabel(self.stats_frame, text="SUCCESS", font=("Inter", 11, "bold"), text_color="#9DABB9").grid(row=1, column=0)

        # Failed Column
        self.failed_val = customtkinter.CTkLabel(self.stats_frame, text="0", font=("Inter", 24, "bold"), text_color="#EF4444")
        self.failed_val.grid(row=0, column=1)
        customtkinter.CTkLabel(self.stats_frame, text="FAILED", font=("Inter", 11, "bold"), text_color="#9DABB9").grid(row=1, column=1)

        # Pending Column
        self.pending_val = customtkinter.CTkLabel(self.stats_frame, text="0", font=("Inter", 24, "bold"), text_color="#FFFFFF")
        self.pending_val.grid(row=0, column=2)
        customtkinter.CTkLabel(self.stats_frame, text="PENDING", font=("Inter", 11, "bold"), text_color="#9DABB9").grid(row=1, column=2)

        # --- SEND BUTTON ---
        # Upgraded to CTkButton to match your theme
        btn_image = Image.open("images/send_message.png")
        
        btn_photo_image = customtkinter.CTkImage(light_image=btn_image, size=(50,50))
        self.send_btn = customtkinter.CTkButton(
            self.scrollbar, text="Launch Campaign", command=self.start_sending, 
            height=50, width=300, font=("Inter", 16, "bold"), fg_color="#3B82F6", 
            hover_color="#2563EB", image=btn_photo_image, compound='left'
        )
        self.send_btn.pack(fill="x", padx=40, pady=50)

        # State Variables
        self.number_of_phones = 0
        self.number_of_phones_done = 0


      
    def open_excel_frame(self,event=None):
       if self.file_path:
        os.startfile(self.file_path)
    
    def make_file_card_clickable(self):
        """Make file card clickable with hover effect"""

        def on_enter(event=None):
            self.file_card.configure(fg_color="#252B33")

        def on_leave(event=None):
            self.file_card.configure(fg_color="#1C2127")

        def bind_hover(widget):
            # 1. Check if the widget is the delete button OR a child of the delete button
            # We use str() comparison because winfo_parent returns a string path
            if widget == self.delete_btn or str(self.delete_btn) in str(widget):
                return # Skip everything related to the delete button

            # 2. Bind hover effects
            widget.bind("<Enter>", on_enter)
            widget.bind("<Leave>", on_leave)
            widget.configure(cursor="hand2")
            
            # 3. Bind the click event to open the file
            widget.bind("<Button-1>", self.open_excel_frame)

        # Bind frame and all children recursively
        bind_hover(self.file_card)
        for child in self.file_card.winfo_children():
            bind_hover(child)
            if hasattr(child, "winfo_children"):
                for grandchild in child.winfo_children():
                    bind_hover(grandchild)

        self.delete_btn.configure(cursor="arrow")
        # Stop delete button from triggering card click
        


    def upload_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        if file_path:
            self.file_path = file_path
            
            # Get File Stats
            file_size = os.path.getsize(file_path) / 1024 # KB
            try:
                df = pd.read_excel(file_path)
                contact_count = len(df)
            except Exception:
                contact_count = 0

            # Update Labels
            self.file_name_label.configure(text=os.path.basename(file_path))
            self.file_stats_label.configure(text=f"{file_size:.1f} KB • {contact_count} Contacts loaded")
            
            # UI Swap: Hide upload button, show file card
            self.file_card.pack(fill="x", padx=40, pady=20, before=self.delay_message)

    

    def remove_file(self):
        self.file_path = None
        self.file_card.pack_forget()

        
    def start_sending(self):
        if not self.file_path:
            messagebox.showerror("Error", "Please upload an Excel file first.")
            return

        # Run sending in separate thread (so GUI doesn't freeze)
        messagebox.showinfo("Pending", "The messages under sending")
        threading.Thread(target=self.send_messages, daemon=True).start()

    def update_progress(self, success_count, failed_count, total):
        done = success_count + failed_count
        percentage = done / total
        
        # Update the Bar
        self.progressbar.set(percentage)
        
        # Update the Texts
        self.percent_label.configure(text=f"{int(percentage * 100)}%")
        self.counter_label.configure(text=f"Messages Sent: {done} of {total}")
        
        # Update the Stats
        self.success_val.configure(text=str(success_count))
        self.failed_val.configure(text=str(failed_count))
        self.pending_val.configure(text=str(total - done))

    def initialize_progress_ui(self):
        # Show the whole dashboard container
        self.progress_container.pack(fill="x", padx=40, pady=20, before=self.send_btn)
        
        # Reset visual state
        self.progressbar.set(0)
        self.success_val.configure(text="0")
        self.failed_val.configure(text="0")
        self.pending_val.configure(text=str(self.number_of_phones))
        self.percent_label.configure(text="0%")
        self.counter_label.configure(text=f"Messages Sent: 0 of {self.number_of_phones}")

    def restart_progress_ui(self):
        # Hide the dashboard
        self.progress_container.pack_forget()

    def send_messages(self):
        try:
            wb = load_workbook(self.file_path)
            ws = wb.active
            
            report_wb = Workbook()
            report_ws = report_wb.active
            report_ws.title = "Report"
            report_ws.append(["Phone", "State", "Date"])

            self.number_of_phones = len(ws['A']) - 1
            self.number_of_phones_done = 0
            success_count = 0
            failed_count = 0

            self.root.after(0, self.initialize_progress_ui)

            bot = Auto_Whatsapp(msg="", people="")
            bot.ensure_logged_in()

            if not self.delay_input.get() or int(self.delay_input.get().strip()) < 0:
                messagebox.showerror("Error", "You must specifiy a delay number or you are going to get banned!")
                self.restart_progress_ui()
                bot.quit_from_process()
                return


            DELAY_BETWEEN_MESSAGES = int(self.delay_input.get().strip())

            

            for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
                phone, message = row

                if phone and message:
                    
                    formatted_phone = bot.normalize_phone(phone)

                    if not formatted_phone:
                        state = "INVALID NUMBER ❌"
                        failed_count += 1
                        self.number_of_phones_done += 1
                        temp_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                        report_ws.append([phone, state, temp_date])
                        continue

                    bot.msg = str(message).strip()
                    bot.target = formatted_phone


                    try:
                        state = bot.send_message(formatted_phone, bot.msg)
                        if "SENT ✅" in state:
                            success_count += 1
                        else:
                            failed_count += 1
                    except Exception as e:
                        print("Error:", e)
                        state = "FAILED ❌"
                        failed_count += 1

                    # Move the increment and UI update here, once per loop
                    self.number_of_phones_done += 1
                    self.root.after(0, lambda s=success_count, f=failed_count, t=self.number_of_phones: 
                                    self.update_progress(s, f, t))
                    
                    temp_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    report_ws.append([phone, state, temp_date])

                    sleep_time = DELAY_BETWEEN_MESSAGES + random.randint(2, 6)
                    print(f"Waiting {sleep_time} seconds...\n")
                    time.sleep(sleep_time)
            self.restart_progress_ui()
            report_wb.save("report.xlsx")
            messagebox.showinfo("Done", "All messages sent!\nReport saved as report.xlsx")

        except Exception as e:
            messagebox.showerror("Error", str(e))


# =========================
# RUN GUI
# =========================
def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

myappid = "mycompany.whatsappautomation.app"
ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid)
if __name__ == "__main__":
    root = tk.Tk()
    root.geometry("1000x600")
    root.configure(bg="#101922")
    icon_path = resource_path("icon.ico")
    if os.path.exists(icon_path):
        root.iconbitmap(icon_path)
    app = WhatsAppGUI(root)
    root.mainloop()

# pyinstaller --onefile --windowed --icon=icon.ico --add-data "images/*.png:images"  autoWhatsapp.py
#datas=[('icon.ico', '.'), ('logo_excel.png', '.'), ('wall-clock.png', '.'), ('send_message.png', '.')],
#hiddenimports=['selenium.webdriver.chrome.options', 'selenium.webdriver.chrome.webdriver'],
